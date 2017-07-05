import os
import openpyxl as xl
import json
import pandas as pd

# This should only contain xlsx files!
raw_directory = os.path.join('data', 'raw')
intermediate_directory = os.path.join('data','intermediate')
meta_directory = os.path.join('data','meta')

def register_template(xlsx_template_name):
    """
    This function reads in a master xlsx file and will record all of the sheet
    and column names, excluding those which start with a '#', as they're treated
    as documentation sheets/columns.

    The result is serialised as JSON and the dict is returned.
    """
    template_workbook = xl.load_workbook(os.path.join(meta_directory,xlsx_template_name), data_only=True)

    sheet_column_dict = {}

    for sheetname in template_workbook.sheetnames:
        # Exclude comment sheets
        if sheetname[0] is '#':
            pass
        else:
            sheet = template_workbook.get_sheet_by_name(sheetname)
            columns = [cell.value for cell in sheet[1] if cell.value != None and cell.value[0] != '#']
            sheet_column_dict[sheetname] = columns

    with open(os.path.join(meta_directory, 'template_column_map.json'), 'w') as f:
        f.write(json.dumps(sheet_column_dict, indent='  '))

    return sheet_column_dict

def create_wb_dict(xlsx_directory):
    '''
    Create a dictionary of all the workbooks in the raw directory, and puts them
    into a dict where the file name is the key and the whole workbook is the value.

    Param: xlsx_directory - this should only contain xlsx files
    '''
    temp_wb_dict = {}
    for workbook_path in os.listdir(xlsx_directory):
        temp_wb_dict[workbook_path] = xl.load_workbook(os.path.join(xlsx_directory,workbook_path), data_only=True)
    return temp_wb_dict

def get_file_names_mapping(path_to_file_map_json, workbook_dict):
    """
    Takes a path to the filename mapping and retuns a dictionary which can be
    queried so that a file name can yeild an identifier for respondents
    (currently a country).

    It asserts that the filenames provided match those found in the 'raw' directory.
    """
    temp_file_map = json.load(open(path_to_file_map_json,'r'))
    # use set comparison to make sure there's no discrepancy between our mapping file and our directory
    assert(set(temp_file_map.keys()) == set(workbook_dict.keys()))
    return temp_file_map

def get_alterations(path_to_alterations_file):
    """
    gets a structured list of alterations to apply to each workbook
    """
    return json.load(open(path_to_alterations_file,'r'))


# 2. Diagnostics

# 2.1 Check if all of the workbooks have the same sheetnames...
def get_non_standard_sheetnames(workbook_dict, predefined_list=[]):
    # Get the list of sheetname for each workbook and store them in a list
    temp_sheetnames = [book.sheetnames for book in workbook_dict.values()]
    # Make a set of sheetnames included in *all* of the workbooks, unless a user
    # defined list has been provided
    if len(predefined_list) < 1:
        intersection_set = set(temp_sheetnames[0]).intersection(*temp_sheetnames[1:])
    else:
        intersection_set = set(predefined_list)
    # Make a set of sheetnames included in *any* of the workbooks
    union_set = set(temp_sheetnames[0]).union(*temp_sheetnames[1:])
    # Return all of the
    return union_set.difference(intersection_set)

def delete_unwanted_sheets(workbook_dict, delete_list):
    '''
    Given a dict of workbooks and a list of sheets, this methog will iterate
    thought all of the books and delete any sheets which are found in the
    delete-list
    '''
    for book in workbook_dict.values():
        for sheet in book.sheetnames:
            if sheet in delete_list:
                book.remove_sheet(book.get_sheet_by_name(sheet))
    return workbook_dict

def make_heading_substitutions(workbook_dict, alteration_dict):
    for name, book in workbook_dict.items():
        for sheet in book:
            for cell in sheet['1']:
                if cell.value in alteration_dict['substitutions']['columns'].keys():
                    old_val = cell.value
                    cell.value = alteration_dict['substitutions']['columns'][cell.value]
                    print('value corrected\n---{}\n:::{}'.format(old_val,cell.value))

def clean_suspected_heading_duplicates(workbook_dict, alteration_dict):
    for name, book in workbook_dict.items():
        for sheetname in book.sheetnames:
            # for each sheet which has a suspected duplication
            if sheetname in alteration_dict['column_deduplications'].keys():
                sheet = book.get_sheet_by_name(sheetname)
                # for each suspected duplication in the sheet
                for col_letter, duplicate_val in alteration_dict['column_deduplications'][sheetname].items():
                    # confirm the duplicate value
                    if sheet[col_letter + '1'].value == duplicate_val:
                        print("Deleting:{}:\n\t{}:\n\t\t{}: {}".format(name, sheetname, col_letter, duplicate_val))
                        for cell in sheet[col_letter]:
                            cell.value = None

# 2.2 Check the column names for each sheet in the same way as the above
#

# 3. Transformation


# Add the country id to each sheet in each workbook
# for filename, wb in wb_dict.items():
#     country = file_map[filename]



def save_wb_dict(workbook_dict, output_directory):
    """
    DON'T TRY TO SAVE OVER FILES - OPENPYXL IS A BELLEND ABOUT THIS AND FAILS
    WITHOUT EXPLANATION

    This is just a helper method for saving a batch of workbooks to a *new*
    directory
    """
    for name, workbook in workbook_dict.items():
        workbook.template = False
        workbook.save(os.path.join(output_directory, name))

def get_column_counts(workbook_dict):
    """
    This expects workbook dictionaries that have been cleaned i.e. have
    consistent sheet names throughout. It returns a mapping from file ids to
    the number of columns in its workbook.
    """
    temp_column_name_map = {}
    for name, workbook in workbook_dict.items():
        temp_column_name_map[name] = []
        for sheet in workbook:
            temp_column_name_map[name].append(len(sheet['1']))

    return temp_column_name_map

def get_consolidated_workbook(workbook_dict, template_dict, data_starting_row=3):
    """
    Because openpyxl is horrifying, this method will consolidate a dictionary
    of workbooks into a single dictionary of pandas dataframes, each a
    consolidation of a sheet and its expected columns as found in the dictionary
    of openpyxl worbooks.

    This method takes your dictionary of workbooks, the template specification
    for a workbook, and the row the actual data starts at, presuming it has
    proper headings and several documentation rows beneath them.

    Remember: the data_starting_row count starts a 1 not 0 because openpyxl.
    """
    # for every sheet in the extracted template
    pandas_sheet_dict = {}
    for template_sheetname, template_column_list in template_dict.items():
        # create a dataframe with the same column names
        pandas_sheet = pd.DataFrame(columns=template_column_list)
        # for every workbook in the dict
        for name, workbook in workbook_dict.items():
            # create a mini data frame for this workbook's specific sheet
            single_pandas_sheet = pd.DataFrame(columns=template_column_list)
            # get the relevant sheet from the workbook
            workbook_sheet = workbook.get_sheet_by_name(template_sheetname)
            # iterate over the columns
            for column in workbook_sheet.columns:
                # if the column is in the template, then set the column in the
                # pandas data frame to the values in the workbook sheet
                if column[0].value in template_column_list:
                    single_pandas_sheet[column[0].value] = [cell.value for cell in column[data_starting_row:]]

            pandas_sheet = pandas_sheet.append(single_pandas_sheet, ignore_index = False)

        # clean out na values
        pandas_sheet = pandas_sheet[pandas_sheet['iati-identifier'].notnull()].dropna(how='all')
        # if there's no dataframe in the dict, log this one
        try:
            pandas_sheet_dict[template_sheetname] = pandas_sheet_dict[template_sheetname].append(pandas_sheet, ignore_index)
        except KeyError:
            pandas_sheet_dict[template_sheetname] = pandas_sheet

    return pandas_sheet_dict

def pandas_dict_to_excel(pandas_sheet_dict, output_path):
    writer = pd.ExcelWriter(output_path)
    for sheetname, dataframe in pandas_sheet_dict.items():
        dataframe.to_excel(writer, sheetname)
    writer.save

##########
# Script #
##########

# Create the dictionary of workbooks
# wb_dict = create_wb_dict(raw_directory)

# get the expected structure from a live template NOTE: you could also write a json file for this...
template_structure = register_template('ActionAid-Template.xlsx')

# Create the sheetname list from json meta-data
sheetnames = template_structure.keys()

# Get and varify the file map from json meta-data
file_map = get_file_names_mapping(os.path.join('data','meta','file-mapping.json'), wb_dict)

# Find all the sheets which should be purged
sheets_to_delete = get_non_standard_sheetnames(wb_dict, sheetnames)

# Purge them all
delete_unwanted_sheets(wb_dict, sheets_to_delete)

# save a backup somewhere sensible in the intermediate_directory
# save_wb_dict(wb_dict, os.path.join(intermediate_directory, 'archive', 'extra_sheets_removed'))

alterations = get_alterations(os.path.join('data','meta', 'alterations.json'))

# change any heading names which are wrong
make_heading_substitutions(wb_dict, alterations)

# save_wb_dict(wb_dict, os.path.join(intermediate_directory, 'archive', 'extra_sheets_removed_plus_multiple_colnames_corrected'))

# delete any suspected duplicate heading locations
clean_suspected_heading_duplicates(wb_dict, alterations)

# save somewhere sensible in the intermediate_directory
# save_wb_dict(wb_dict, os.path.join(intermediate_directory,'archive','corrections_and_deduplication'))

get in the new batch
# wb_dict_deduplicated = create_wb_dict(os.path.join(intermediate_directory,'archive','corrections_and_deduplication'))

cons_wb_df = get_consolidated_workbook(wb_dict_deduplicated, template_structure)

pandas_dict_to_excel(cons_wb_df, os.path.join(intermediate_directory, 'consolidated_workbook', 'output.xlsx'))

############
# Drafting #
############

temp_wb = xl.Workbook()

wb1 = xl.load_workbook('AAN 2016 IATI Report - Nigeria.xlsx', data_only=True)

wb1_activities = wb1.get_sheet_by_name('Activity Level')

for column in wb1_activities.columns:
    if column[0].value == 'activity-date/0/@iso-date':
        print([cell.value for cell in column[3:]])

wb1.template = False
wb1.save(os.path.join('modified___AAN 2016 IATI Report - Nigeria.xlsx'))
